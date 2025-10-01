import streamlit as st
import pandas as pd
import io
import zipfile
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Настройка страницы
st.set_page_config(
    page_title="Переоценка квартир",
    page_icon="🏡",
    layout="wide"
)

# Заголовок с иконкой
col1, col2 = st.columns([1, 10])
with col1:
    st.image("my_icon.png", width=50)
with col2:
    st.markdown("## Переоценка квартир по подразделениям")

st.markdown("""
Загрузите Excel с квартирами, выберите фильтры и сумму прибавки.  
Приложение покажет превью данных и создаст архив с переоценёнными файлами.
""")

# Загрузка файла
uploaded_file = st.file_uploader("📥 Загрузите Excel-файл", type=["xlsx"])

if uploaded_file:
    try:
        df = pd.read_excel(uploaded_file)
    except Exception as e:
        st.error(f"Ошибка при чтении файла: {e}")
    else:
        required_cols = {"Готовность объекта", "Подразделение", "Номер квартиры", "Этаж",
                         "Площадь общая", "Тип квартиры", "Статус", "Вид помещения", "Стоимость"}
        if not required_cols.issubset(df.columns):
            st.error(f"Файл должен содержать колонки: {', '.join(required_cols)}")
        else:
            # Переименуем колонку "Стоимость"
            df = df.rename(columns={"Стоимость": "Стоимость по пред.приказу"})

            # --- Фильтры с "Выбрать всё" ---
            def multiselect_all(label, options):
                opts = options.tolist()
                opts.insert(0, "Все")  # добавляем опцию "Все"
                selected = st.multiselect(label, opts)
                if "Все" in selected or not selected:
                    return options  # выбираем все
                return selected

            # Фильтр по готовности
            readiness_choices = multiselect_all("Выберите готовность объекта:", df["Готовность объекта"].unique())
            df_filtered = df[df["Готовность объекта"].isin(readiness_choices)]

            # Фильтр по подразделениям
            dept_choices = multiselect_all("Выберите подразделения для переоценки:", df_filtered["Подразделение"].unique())
            df_filtered = df_filtered[df_filtered["Подразделение"].isin(dept_choices)]

            # Фильтр по статусу
            status_choices = multiselect_all("Выберите статус:", df_filtered["Статус"].unique())
            df_filtered = df_filtered[df_filtered["Статус"].isin(status_choices)]

            # Фильтр по виду помещения
            vid_choices = multiselect_all("Выберите вид помещения:", df_filtered["Вид помещения"].unique())
            df_filtered = df_filtered[df_filtered["Вид помещения"].isin(vid_choices)]

            # Фильтр по типу квартиры
            flat_type_choices = multiselect_all("Выберите тип квартиры:", df_filtered["Тип квартиры"].unique())
            df_filtered = df_filtered[df_filtered["Тип квартиры"].isin(flat_type_choices)]

            # Сумма прибавки
            add_val = st.number_input("Сколько добавить к стоимости (₽):", step=10000, min_value=0)

            # Дата для имени файла
            report_date = st.date_input("📅 Выберите дату для имени файла")

            # Превью данных
            if not df_filtered.empty:
                preview_df = df_filtered.copy()
                preview_df["Новая стоимость"] = preview_df["Стоимость по пред.приказу"] + add_val
                preview_df["Новая цена кв.м"] = preview_df["Новая стоимость"] / preview_df["Площадь общая"]
                preview_df["Изменение"] = preview_df["Новая стоимость"] - preview_df["Стоимость по пред.приказу"]

                preview_df = preview_df[[
                    "Номер квартиры", "Этаж", "Площадь общая", "Тип квартиры",
                    "Статус", "Вид помещения", "Стоимость по пред.приказу", "Новая стоимость",
                    "Новая цена кв.м", "Изменение"
                ]]

                # Итоговая строка
                totals = {
                    "Номер квартиры": "ИТОГО:",
                    "Стоимость по пред.приказу": preview_df["Стоимость по пред.приказу"].sum(),
                    "Новая стоимость": preview_df["Новая стоимость"].sum(),
                    "Изменение": preview_df["Изменение"].sum()
                }
                totals_df = pd.DataFrame([totals])
                preview_with_total = pd.concat([preview_df, totals_df], ignore_index=True)

                st.subheader("Превью таблицы с новой стоимостью")
                st.dataframe(preview_with_total)

            # Генерация ZIP с Excel
            if st.button("Выполнить пересчёт"):
                if df_filtered.empty:
                    st.warning("Нет данных для пересчёта по выбранным фильтрам!")
                else:
                    buffer = io.BytesIO()
                    with zipfile.ZipFile(buffer, "w") as zf:
                        date_str = report_date.strftime("%d.%m.%Y")

                        for dept in df_filtered["Подразделение"].unique():
                            df_dept = df_filtered[df_filtered["Подразделение"] == dept].copy()
                            df_dept["Новая стоимость"] = df_dept["Стоимость по пред.приказу"] + add_val
                            df_dept["Новая цена кв.м"] = df_dept["Новая стоимость"] / df_dept["Площадь общая"]
                            df_dept["Изменение"] = df_dept["Новая стоимость"] - df_dept["Стоимость по пред.приказу"]

                            df_dept = df_dept[[
                                "Номер квартиры", "Этаж", "Площадь общая", "Тип квартиры",
                                "Статус", "Вид помещения", "Стоимость по пред.приказу", "Новая стоимость",
                                "Новая цена кв.м", "Изменение"
                            ]]

                            totals = {
                                "Номер квартиры": "ИТОГО:",
                                "Стоимость по пред.приказу": df_dept["Стоимость по пред.приказу"].sum(),
                                "Новая стоимость": df_dept["Новая стоимость"].sum(),
                                "Изменение": df_dept["Изменение"].sum()
                            }
                            totals_df = pd.DataFrame([totals])
                            df_dept = pd.concat([df_dept, totals_df], ignore_index=True)

                            # Создание Excel с оформлением
                            wb = Workbook()
                            ws = wb.active

                            for r in dataframe_to_rows(df_dept, index=False, header=True):
                                ws.append(r)

                            # --- Стили ---
                            thin = Side(border_style="thin", color="FFFFFF")  # белые границы
                            border = Border(left=thin, right=thin, top=thin, bottom=thin)
                            header_fill = PatternFill("solid", fgColor="F79646")
                            zebra_fill = PatternFill("solid", fgColor="D9D9D9")
                            header_font = Font(name="Calibri Light", size=9, bold=True)
                            cell_font = Font(name="Calibri Light", size=9, bold=False)
                            align_center = Alignment(horizontal="center", vertical="center")

                            max_row = ws.max_row
                            max_col = ws.max_column

                            # Заголовки
                            for cell in ws[1]:
                                cell.fill = header_fill
                                cell.font = header_font
                                cell.alignment = align_center
                                cell.border = border

                            # Тело таблицы
                            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=max_row-1, min_col=1, max_col=max_col), start=1):
                                for cell in row:
                                    cell.font = cell_font
                                    cell.border = border
                                    cell.alignment = align_center
                                    if row_idx % 2 == 0:
                                        cell.fill = zebra_fill
                                    if cell.column == 3:
                                        cell.number_format = '#,##0.00'
                                    elif isinstance(cell.value, (int, float)):
                                        cell.number_format = '#,##0'

                            # Итоговая строка
                            for cell in ws[max_row]:
                                cell.fill = header_fill
                                cell.font = header_font
                                cell.alignment = align_center
                                cell.border = border
                                if cell.column == 3:
                                    cell.number_format = '#,##0.00'
                                elif isinstance(cell.value, (int, float)):
                                    cell.number_format = '#,##0'

                            # Автоширина колонок
                            for col in ws.columns:
                                max_length = 0
                                col_letter = col[0].column_letter
                                for cell in col:
                                    try:
                                        if cell.value:
                                            max_length = max(max_length, len(str(cell.value)))
                                    except:
                                        pass
                                ws.column_dimensions[col_letter].width = max_length + 2

                            final_buffer = io.BytesIO()
                            wb.save(final_buffer)
                            final_buffer.seek(0)
                            zf.writestr(f"{dept}_{date_str}.xlsx", final_buffer.getvalue())

                    buffer.seek(0)
                    st.success("Файлы готовы! Скачайте архив ниже.")
                    st.download_button(
                        label="📥 Скачать архив",
                        data=buffer,
                        file_name="recalculated_departments.zip",
                        mime="application/zip"
                    )